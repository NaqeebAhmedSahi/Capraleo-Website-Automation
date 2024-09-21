import undetected_chromedriver as uc
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from bs4 import BeautifulSoup
import requests
import os
import time
from PIL import Image
from io import BytesIO
import openpyxl
from openpyxl.drawing.image import Image as OpenpyxlImage

# Base URL
base_url = "https://capraleo.com/"

def fetch_category_links(driver):
    """Fetch all category links from the specified class."""
    try:
        driver.get(base_url)
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CLASS_NAME, "ekit_badge_left")))
        soup = BeautifulSoup(driver.page_source, 'html.parser')
        category_elements = soup.find_all('a', class_='ekit_badge_left')
        category_links = [a['href'] for a in category_elements if 'href' in a.attrs]
        print(f"Found category links: {category_links}")
        return category_links
    except Exception as e:
        print(f"Error fetching category links: {e}")
        return []

def fetch_sub_category_links(driver, category_link):
    """Fetch all subcategory links under a given category."""
    try:
        driver.get(category_link)
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CLASS_NAME, "elementor-container")))
        soup = BeautifulSoup(driver.page_source, 'html.parser')
        sub_category_elements = soup.find_all('a', href=True)
        sub_category_links = [a['href'] for a in sub_category_elements if "category" in a['href']]
        print(f"Found subcategory links for {category_link}: {sub_category_links}")
        return sub_category_links
    except Exception as e:
        print(f"Error fetching subcategory links from {category_link}: {e}")
        return []

def read_scraped_links(log_file_path):
    """Read already scraped links from a log file."""
    if os.path.exists(log_file_path):
        with open(log_file_path, 'r') as file:
            return set(line.strip() for line in file)
    return set()

def log_scraped_link(link, log_file_path):
    """Log the scraped link to avoid reprocessing."""
    try:
        with open(log_file_path, 'a') as file:
            file.write(link + '\n')
    except Exception as e:
        print(f"Error logging scraped link {link}: {e}")

def download_and_resize_image(image_url, image_name, max_size=(100, 100)):
    """Download and resize the image."""
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
    }
    try:
        response = requests.get(image_url, headers=headers, stream=True)
        response.raise_for_status()
        img = Image.open(BytesIO(response.content))
        img.thumbnail(max_size)  # Resize the image
        img_path = os.path.join(image_folder, image_name)
        img.save(img_path)
        return img_path
    except Exception as e:
        print(f"Failed to download or resize image from {image_url}. Error: {e}")
        return 'N/A'

def scrape_product_data(driver, sheet, row):
    """Scrape product data from the webpage and write to Excel."""
    try:
        soup = BeautifulSoup(driver.page_source, 'html.parser')

        # Extract data using BeautifulSoup
        category = soup.find(class_='single-product-category')
        title = soup.find(class_='product_title entry-title')
        price_element = soup.find(class_='woocommerce-Price-amount')
        sku_wrapper = soup.find(class_='sku_wrapper')
        sku = soup.find(class_='sku')
        description_div = soup.find('div', id='tab-description')
        size = soup.find(class_='woocommerce-product-attributes-item__value')
        image_tag = soup.find('img', class_='wp-post-image')  # Extract image with this class

        # Extract text with default values if elements are not found
        category_text = category.get_text(strip=True) if category else 'N/A'
        title_text = title.get_text(strip=True) if title else 'N/A'
        
        if price_element:
            price_currency = price_element.find(class_='woocommerce-Price-currencySymbol').get_text(strip=True) if price_element.find(class_='woocommerce-Price-currencySymbol') else 'N/A'
            price_value = price_element.find('bdi').get_text(strip=True) if price_element.find('bdi') else 'N/A'
            price_text = f"{price_currency} {price_value}"
        else:
            price_text = 'N/A'
        
        sku_combined = f"{sku_wrapper.get_text(strip=True)} {sku.get_text(strip=True)}" if sku_wrapper and sku else 'N/A'
        description_text = description_div.find('p').get_text(strip=True) if description_div and description_div.find('p') else 'N/A'
        size_text = size.get_text(strip=True) if size else 'N/A'
        
        # Extract image URL and download the image
        if image_tag:
            image_url = image_tag['src']
            image_name = image_url.split('/')[-1]
            image_path = download_and_resize_image(image_url, image_name)
        else:
            image_path = 'N/A'

        # Write the data to the Excel file
        sheet.cell(row=row, column=1, value=category_text)
        sheet.cell(row=row, column=2, value=title_text)
        sheet.cell(row=row, column=3, value=price_text)
        sheet.cell(row=row, column=4, value=sku_combined)
        sheet.cell(row=row, column=5, value=description_text)
        sheet.cell(row=row, column=6, value=size_text)

        if image_path != 'N/A':
            img = OpenpyxlImage(image_path)
            sheet.add_image(img, f"G{row}")
        print(f"Data saved in Excel: Category: {category_text}, Title: {title_text}, Price: {price_text}, SKU Combined: {sku_combined}, Description: {description_text}, Size: {size_text}, Image: {image_path}")
        
    except Exception as e:
        print(f"Failed to scrape product data. Error: {e}")

def process_sub_category(driver, sub_category_link, sub_category_name):
    """Process each subcategory by scraping its products."""
    print(f"Processing subcategory: {sub_category_name}")

    # File paths for the current subcategory
    log_file_path = f'{sub_category_name}_scraped_links.txt'
    excel_file_path = f'{sub_category_name}_product_data.xlsx'
    global image_folder
    image_folder = f'images/{sub_category_name}'  # Unique image folder for each subcategory

    # Ensure the image folder and necessary directories exist
    if not os.path.exists(image_folder):
        os.makedirs(image_folder)

    # Ensure the log file path and Excel file path directories exist
    log_file_dir = os.path.dirname(log_file_path)
    if log_file_dir and not os.path.exists(log_file_dir):
        os.makedirs(log_file_dir)

    excel_file_dir = os.path.dirname(excel_file_path)
    if excel_file_dir and not os.path.exists(excel_file_dir):
        os.makedirs(excel_file_dir)

    scraped_links = read_scraped_links(log_file_path)

    # Prepare Excel file
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = f"{sub_category_name} Data"
    sheet.append(['Category', 'Title', 'Price', 'SKU Combined', 'Description', 'Size', 'Image'])
    sheet.column_dimensions['G'].width = 20  # Adjust the width as needed

    row = 2

    page_number = 1
    while True:
        page_url = f"{sub_category_link}/page/{page_number}/"
        new_page_url =  page_url
        driver.get(page_url)
        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CLASS_NAME, "products.columns-4")))

        li_elements = driver.find_elements(By.CSS_SELECTOR, ".products.columns-4 li")
        product_links = [li.find_element(By.TAG_NAME, 'a').get_attribute('href') for li in li_elements if li.find_element(By.TAG_NAME, 'a')]

        if not product_links:
            print(f"No product links found on page {page_number} for subcategory: {sub_category_link}")
            break  # Exit loop if no products are found on this page
        else:
            print(f"Found {len(product_links)} product links on page {page_number} for subcategory: {sub_category_link}")

        for link in product_links:
            if link in scraped_links:
                print(f"Link already scraped: {link}")
                continue
            
            try:
                print(f"Processing product link: {link}")
                driver.get(link)
                WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.TAG_NAME, "body")))
                print(f"Opened product page: {link}")
                
                scrape_product_data(driver, sheet, row)
                row += 1
                log_scraped_link(link, log_file_path)
                time.sleep(1)  # To avoid being flagged for scraping too quickly
                
            except Exception as e:
                print(f"Error processing link {link}: {e}")
                continue
       

        # Check if there's a next page
        try:
            # Redirect to the main subcategory page to check for additional pages
            driver.get(new_page_url)
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CLASS_NAME, "products.columns-4")))
            soup = BeautifulSoup(driver.page_source, 'html.parser')
            next_page = soup.find('a', class_='next page-numbers')
            if not next_page:
                print(f"No more pages found for subcategory: {sub_category_link}")
                break  # Exit loop if no more pages are found
            else:
                print(f"Moving to next page for subcategory: {sub_category_link}")
                page_number += 1
        except Exception as e:
            print(f"Error checking for additional pages: {e}")
            break  # Exit loop if there's an error checking for next page

    workbook.save(excel_file_path)
    print(f"Data saved to Excel file: {excel_file_path}")

try:
    options = uc.ChromeOptions()
    options.add_argument("--headless=new")  # Enable new headless mode
    options.add_argument("--disable-gpu")  # Disable GPU usage
    options.add_argument("--disable-extensions")  # Disable extensions
    options.add_argument("--no-sandbox")  # Disable sandbox mode for security
    options.add_argument("--disable-dev-shm-usage")  # Prevents issues on systems with low memory
    options.add_argument("--disable-blink-features=AutomationControlled")  # Avoid detection as automated

    with uc.Chrome(options=options) as driver:
        print("Driver initialized successfully in headless mode.")
        category_links = fetch_category_links(driver)
        
        if not category_links:
            print("No category links found. Exiting.")
        else:
            for sub_category_link in category_links:  # Using category_links directly
                sub_category_name = sub_category_link.split('/')[-2]
                process_sub_category(driver, sub_category_link, sub_category_name)

except Exception as e:
    print(f"An error occurred: {e}")

finally:
    try:
        driver.quit()
    except Exception as e:
        print(f"Failed to quit driver: {e}")
